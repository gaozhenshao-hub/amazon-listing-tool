from collections import Counter
from pathlib import Path
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/产品表现ASIN（2026-08-03~2026-08-09，全部广告）-946369827630383104.xlsx')
workbook = load_workbook(source, read_only=True, data_only=False)

for worksheet in workbook.worksheets:
    rows = worksheet.iter_rows(values_only=True)
    header = list(next(rows))
    samples = []
    columns = [[ ] for _ in header]

    for row_number, row in enumerate(rows, start=2):
        padded_row = list(row) + [None] * max(0, len(header) - len(row))
        if len(samples) < 5:
            samples.append((row_number, padded_row))
        for index, value in enumerate(padded_row[:len(header)]):
            if value not in (None, ''):
                columns[index].append(value)

    print(f'工作表：{worksheet.title}')
    print(f'尺寸：{worksheet.max_row} 行 × {worksheet.max_column} 列')
    for index, label in enumerate(header, start=1):
        print(f'列{index}: {label}')

    print('前 5 条非空样例：')
    for row_number, row in samples:
        values = [f'{header[i]}={value}' for i, value in enumerate(row[:len(header)]) if value not in (None, '')]
        print(f'行{row_number}: ' + ' | '.join(values))

    print('关键字段非空率与样例计数：')
    targets = ['日期', 'ASIN', '父ASIN', 'SKU', '店铺', '站点', '可售', '在途', '库存', '销量', '订单', '本地']
    reported_indexes = set()
    for target in targets:
        for index, label in enumerate(header):
            if index in reported_indexes or not label or target.lower() not in str(label).lower():
                continue
            reported_indexes.add(index)
            values = columns[index]
            print(f'{label}: non_empty={len(values)}; samples={values[:3]}')

    print('父 ASIN 与 ASIN 基数：')
    for target in ['父ASIN', 'ASIN']:
        for index, label in enumerate(header):
            if label == target:
                values = [str(value) for value in columns[index]]
                print(f'{target}: distinct={len(set(values))}; top={Counter(values).most_common(5)}')

    date_index = header.index('日期') if '日期' in header else None
    asin_index = header.index('ASIN') if 'ASIN' in header else None
    parent_asin_index = header.index('父ASIN') if '父ASIN' in header else None
    fba_available_index = header.index('FBA-可售') if 'FBA-可售' in header else None
    fba_transit_index = header.index('FBA-在途') if 'FBA-在途' in header else None
    local_available_index = header.index('本地可用') if '本地可用' in header else None
    sales_index = header.index('销量') if '销量' in header else None

    daily_rows = Counter()
    daily_asins: dict[str, set[str]] = {}
    daily_parents: dict[str, set[str]] = {}
    rows_with_positive_inventory = 0
    rows_with_positive_sales = 0
    for row_number, row in samples:
        pass
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        padded_row = list(row) + [None] * max(0, len(header) - len(row))
        if date_index is not None:
            date_value = str(padded_row[date_index])
            daily_rows[date_value] += 1
            if asin_index is not None and padded_row[asin_index] not in (None, ''):
                daily_asins.setdefault(date_value, set()).add(str(padded_row[asin_index]))
            if parent_asin_index is not None and padded_row[parent_asin_index] not in (None, ''):
                daily_parents.setdefault(date_value, set()).add(str(padded_row[parent_asin_index]))
        if fba_available_index is not None and fba_transit_index is not None and local_available_index is not None:
            if sum(int(padded_row[i] or 0) for i in [fba_available_index, fba_transit_index, local_available_index]) > 0:
                rows_with_positive_inventory += 1
        if sales_index is not None and int(padded_row[sales_index] or 0) > 0:
            rows_with_positive_sales += 1

    print('日粒度覆盖：')
    for date_value in sorted(daily_rows):
        print(f'{date_value}: rows={daily_rows[date_value]}; asins={len(daily_asins.get(date_value, set()))}; parents={len(daily_parents.get(date_value, set()))}')
    print(f'正库存行数（FBA可售+FBA在途+本地可用 > 0）：{rows_with_positive_inventory}')
    print(f'正销量行数：{rows_with_positive_sales}')
