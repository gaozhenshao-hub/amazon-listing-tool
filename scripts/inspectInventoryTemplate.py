from pathlib import Path
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/库存规划表模版副本.xlsx')
workbook = load_workbook(source, data_only=False, read_only=True)

for worksheet in workbook.worksheets:
    print(f'工作表：{worksheet.title}')
    print(f'尺寸：{worksheet.max_row} 行 × {worksheet.max_column} 列')
    print('非空单元格（前 40 行）：')
    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(40, worksheet.max_row), values_only=False), start=1):
        values = []
        for cell in row:
            if cell.value not in (None, ''):
                values.append(f'{cell.coordinate}={cell.value}')
        if values:
            print(' | '.join(values))
    print('公式单元格：')
    formula_count = 0
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f'{cell.coordinate}={cell.value}')
                formula_count += 1
                if formula_count >= 120:
                    break
        if formula_count >= 120:
            break
    print(f'公式总数（最多列出前 120 个）：{formula_count}')
